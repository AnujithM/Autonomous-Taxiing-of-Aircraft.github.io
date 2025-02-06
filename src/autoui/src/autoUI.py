#!/usr/bin/env python

import rospy
from geometry_msgs.msg import Twist, Point, Quaternion
import tf
from scipy.optimize import minimize
from math import radians, copysign, sqrt, pow, pi, atan2
from tf.transformations import euler_from_quaternion
from std_msgs.msg import Float32
import numpy as np
import time
import socket
import openpyxl
import os
from scipy import signal

ts = 0.0
tinter = 0.0
tstop = 0.0
start_flag = 0
row = 1
x = 0           #For resetting count and count2 in automatic
instance = 0    #For storing instance in switch from auto to manual in junction
state = 0

wb = openpyxl.Workbook()
sheet = wb.active

with open("Stop.txt",'w') as st:
    st.write("1")
with open("current_flag.txt", 'w') as cf:
    cf.write("0")

rospy.init_node('AutoUI');
velocity_publisher = rospy.Publisher('/cmd_vel', Twist, queue_size=10)

ultrasonic_distance = None
def ultrasonic_distance_callback(data):
    global ultrasonic_distance
    ultrasonic_distance = data.data

rospy.Subscriber('/ultrasonic_distance', Float32, ultrasonic_distance_callback)


def Set_Vel(vx,vw):
    twist = Twist()
    twist.linear.y = 0.0; twist.linear.z = 0.0
    twist.angular.x = 0.0; twist.angular.y = 0.0;
    twist.linear.x = vx
    twist.angular.z = vw

    velocity_publisher.publish(twist)
    time.sleep(0.1)

def getK():
    with open("Auto.txt","r") as f:
        cmd = f.read()
        print(cmd)
    return cmd


d = 1
dp = 1
flag = '1'



count = 0
count2 = 0


def stop_tb():

    global count2
    global count

    count = 0
    count2 = 0
    while True:
        flag_ = 'a'
        with open('current_flag.txt','r') as curr:
            flag_ = curr.read()
        if flag_ == '0':
            print("From stop to manual")
            state = 0
            break
        else:
            #print("In stop")
            Set_Vel(0.0,0.0)

def kill_tb():
    while True:
        flag_ = 'a'
        with open('Stop.txt','r') as curr:
            flag_ = curr.read()
        if flag_ == '0':
            Automatic()
        else:
            #print("In stop")
            Set_Vel(0.0,0.0)
def setState(control):
    try:
        with open("current_flag.txt", 'w') as f:
            f.write(str(control))
    except ValueError:
        print("Value Error occured")
def getState():
    global state
    try:
        with open("current_flag.txt", 'r') as f:
            state = int(f.read())
    except ValueError:
        print("Value Error occured")
def getControl():
    try:
        with open("controls.txt", 'r') as f:
            control = f.read()
        return control
    except ValueError:
        print("Value Error occured")
        return []
def intersection(left):
    t1 = time.time()
    print("At intersection")
    while(time.time() - t1 < 4.1):
        
        Set_Vel(0.12, 0)
    Set_Vel(0, 0)

    t2 = time.time()
    if left:
        t = 5.5
        x = 1
    else:
        t = 2.5
        x = -1
    while(time.time() - t2 < t):
        Set_Vel(0, x*0.2)
    Set_Vel(0,0)
    time.sleep(0.2)

inter = 1
def Automatic():
    global state
    global x
    global inter
    global d
    global count2
    global count
    global start_flag
    global ultrasonic_distance
    kill = 'a'
    print("In Automatic")
    count3 = 0
    # start_flag = start_flag+1
    while True:
        # time.sleep(0.08)
        
        getState()
        # with open("current_flag.txt", 'r') as f:
        #   state = int(f.read())
        if not state:
            break
        # if start_flag==1:
        #   ts = time.time()


        #*********************************************************************
        # x = x+1
        # if x%10==0:
        #   # print('x was',x)
        #   x = 0
        #   count = 0
        #   count2 = 0
        #**********************************************************************


        # ch = '1'
        # with open('current_flag.txt','r') as curr:
        #   ch = curr.read()
        # if ch == '0':
        #   state = 0
        #   break
        # print("Automatic")
        data = 'p'
        with open('Lane_Detection.txt','r') as lane:
            data = lane.read()
            # print(data)
            # fullData = data.split()
            # data = fullData[0]
            # stopData = fullData[1]
        # print(data)
        if data == "None":
            # print(count2,"was count2")
            # print("End of the road")
            print("In None========")
            count2 = count2+1
            time.sleep(0.1)
            print(count2)
            if count2>5:
                print("From None=====================================")
                count2 = 0
                setState(0)
                break
            continue
        if data == "Object":
            
            # print(count2,"was count2")
            # print("End of the road")
            print("Object detected")
            count2 = count2+1
            time.sleep(0.1)
            print(count2)
            
            print("Ultrasonic Distance:", ultrasonic_distance)
            
            if count2>5 and ultrasonic_distance is not None and ultrasonic_distance < 25:
                print("Stopped Vehicle")
                count2 = 0
                setState(0)
                break
            

        # 
        #*****************************************************************************
        # elif data == "Take Request1" or data == "Take Request0":
        #   # print(count,"was count")
        #   count = count+1
        #   if count>1:             #Keep it 3
        #       count = 0
        #       if inter:
        #           print(data)
        #           direction = int(data[-1])
                    
        #           intersection(direction)
        #           setState(1)
        #           inter = 0
        #           break
        #       setState(0)
        #       break
                    

        #   continue
        # elif data =="None Left" or data =="None Right":
        #   print(data)
        #   steer_stop(data)
        #***********************************************************************************
        else:
            
            #print("In data")
            #print("Sending steer")
            
            try:
                d = int(data)
                steer()


            except:
                continue


def Manual():
    global state
    global inter
    inter = 1
    kill = "a"
    vx = 0
    wz = 0
    print("In Manual")
    while True:
        # time.sleep(0.08)
        
        # try:
        #   with open('current_flag.txt','r') as f:
        #       state = int(f.read())
        getState()

        if state:
            break
        else:
            
            # with open("controls.txt", 'r') as f:
            #   vxwz = f.read()
            # if len(vxwz) == 0:
            #   print("Value error occured")
            # else:
            #   vx, wz = map(float, vxwz.split())
            vxwz = getControl()
            if len(vxwz):
                vx, wz = map(float, vxwz.split())


            # print(vx, wz)
            Set_Vel(vx, wz)


max_ang = 2.5

Kp = 0.045#0.005
Kd = 0.01#0.01


# # Define the system dynamics matrices A and B
# A = np.array([[0, 1],
#               [0, 0]])
# B = np.array([[0],
#               [1]])

# # Define the state, control input, and prediction horizon
# N = 10  # Prediction horizon

# # Define the cost matrices (Q and R) for the cost function
# Q = np.array([[1, 0],
#               [0, 1]])  # Adjust these values to tune the controller
# R = np.array([[0.1]])  # Adjust this value



# # Define the LQR controller gain matrix K
# Q = np.array([[2, 0],
#               [0, 1.38]])  # Adjust these values to tune the controller
# R = np.array([[0.001]])  # Adjust this value

# # Compute the LQR gain matrix
# poles = np.array([-1, -2])  # Adjust the desired poles
# K = signal.place_poles(A, B, poles).gain_matrix

# # Initialize control signal and previous control signal
# control_signal = 0.0
# previous_control_signal = 0.0

previous_steering_angle = 0.0  # Initialize to 0 or the initial steering angle

def steer():
    global velocity_publisher
    global d
    global dp
    global previous_steering_angle
    # global control_signal
    # global previous_control_signal
    moveL = Twist()
    #print(dp,"was received")
    moveL.linear.x = 0.11
    #PD controller
    if d>0:
        if d*Kp + (d-dp)*Kd> max_ang:
            moveL.angular.z = 2.5
        else:
            moveL.angular.z = d*Kp + (d-dp)*Kd
    else:
        if d*Kp + (d-dp)*Kd<-max_ang:
            moveL.angular.z = -2.94
        else:
            moveL.angular.z = d*Kp + (d-dp)*Kd



    # # SMC
    # K = 2.1 # Controller gain
    # eps = 5 # Threshold
    # error = d - dp

    # # Control signal using sliding mode control
    # if abs(error) <= eps:
    #     control_signal = K * error
    # else:
    #     control_signal = K * eps * copysign(1.0, error)

    # # Limit the control signal to avoid extreme values
    # max_control_signal = 4.0
    # if control_signal > max_control_signal:
    #     control_signal = max_control_signal
    # elif control_signal < -max_control_signal:
    #     control_signal = -max_control_signal
    # #print(moveL.angular.z)
    # moveL.angular.z = control_signal
    # print("value is",d)
    # print(moveL.angular.z)

    # #LQR
    # #Calculate the control input u using LQR controller
    # u = np.dot(K, np.array([[d], [0]]))

    # control_signal = u[0,0]

    # # Apply a simple low-pass filter
    # alpha = 0.02  # Adjust this smoothing factor
    # control_signal = (1 - alpha) * previous_control_signal + alpha * control_signal

    # #Limit the control signal to avoid extreme values
    # max_control_signal = 2.38
    # if control_signal > max_control_signal:
    #     control_signal = max_control_signal
    # elif control_signal < -max_control_signal:
    #     control_signal = -max_control_signal
    # #print(moveL.angular.z)
    # moveL.angular.z = control_signal
    # print(d)
    # print(control_signal)
    
    
    # #Stanley Controller
    # Kp = 0.00092  # Proportional gain
    
    # #heading_error = desired heading angle - actual heading angle
    
    # heading_error = d * 0.0006145013    
    
    # #print(heading_error)
    # steering_angle = heading_error + atan2(Kp * d, 0.1198) # heading error + arctan(Kp * d, desired_speed)

    # angular_velocity = steering_angle/0.14204 #(angular velocity = change in steering angle/r/2*L)
    
    # max_angular_velocity = 2.15
    
    # if angular_velocity > max_angular_velocity:
    #    angular_velocity  = max_angular_velocity
    # elif angular_velocity  < -max_angular_velocity:
    #     angular_velocity  = -max_angular_velocity

    # print(angular_velocity)
    # print("value is", d)
    # moveL.angular.z = angular_velocity
    # moveL.linear.x = 0.121
    moveL.linear.y = 0.0
    moveL.linear.z = 0.0
    moveL.angular.x = 0.0
    moveL.angular.y = 0.0
    velocity_publisher.publish(moveL)
    dp = d
    time.sleep(0.1)

def steer_stop(con):
#    global UDPServerSocket
    global d
    moveL = Twist()
    moveL.linear.x = 0
    moveL.angular.z = 0
    velocity_publisher.publish(moveL)
    if con == "None Right":
        print("Left over steer")
        moveL.linear.x = 0.08
        moveL.angular.z = -0.07
        velocity_publisher.publish(moveL)
    else:
        print("Right over steer")
        moveL.linear.x = 0.08
        moveL.angular.z = 0.07
        velocity_publisher.publish(moveL)
    # data,addr = UDPServerSocket.recvfrom(1024)
    # if data_.decode('utf-8')=="None Left" or data_.decode('utf-8')=="None Right":
    #     time.sleep(0.1)
    #     steer_stop(data_.decode('utf-8'))
    # else:
    #     print(data.decode('utf-8'),"is received")
    #     time.sleep(0.1)
    #     return
    time.sleep(0.1)

# flag = '1'
# ts = time.time()
# c1 = sheet.cell(row = 1, column = 2)
# c1.value = str(ts)
# c2 = sheet.cell(row = 1, column = 1)
# c2.value = "Start Time"
# wb.save("Name.xlsx")

# Automatic()





while(True):
    try:
        if state == 0:
            Manual()
        elif state == 1:
            Automatic()
        elif state == 2:
            stop_tb()
        else:
            print("State is ", state)
            break

    except rospy.ROSInterruptException:
        rospy.loginfo("Ctrl-C caught. Quitting")